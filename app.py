from flask import Flask, render_template, request, redirect, flash
import sqlite3
from datetime import datetime
import os
import pandas as pd

app = Flask(__name__)

app.secret_key = 'chave_secreta_simples'

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
db_path = os.path.join(BASE_DIR, 'banco.db')
excel_path = os.path.join(BASE_DIR, 'analise de dados.xlsx')


def criar_banco():
    print("Banco sendo criado em:", db_path)

    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS amostras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo_servico TEXT,
            data_entrada TEXT,
            data_envio TEXT,
            ensaio1 TEXT,
            ensaio2 TEXT,
            ensaio3 TEXT,
            ensaio4 TEXT
        )
    ''')

    conn.commit()
    conn.close()


criar_banco()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/salvar', methods=['POST'])
def salvar():
    criar_banco()

    codigo_servico = request.form['codigo_servico']
    data_entrada = request.form['data_entrada']
    data_envio = request.form['data_envio']

    ensaio1 = request.form['ensaio1']
    ensaio2 = request.form['ensaio2']
    ensaio3 = request.form['ensaio3']
    ensaio4 = request.form['ensaio4']

    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    c.execute("""
        INSERT INTO amostras 
        (codigo_servico, data_entrada, data_envio, ensaio1, ensaio2, ensaio3, ensaio4)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (codigo_servico, data_entrada, data_envio, ensaio1, ensaio2, ensaio3, ensaio4))

    conn.commit()
    conn.close()

    flash('Amostra cadastrada com sucesso!')
    return redirect('/')


@app.route('/lista')
def lista():
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    c.execute("SELECT * FROM amostras")
    dados = c.fetchall()

    conn.close()

    resultados = []

    for d in dados:
        try:
            data_entrada = datetime.strptime(d[2], "%Y-%m-%d")
            data_envio = datetime.strptime(d[3], "%Y-%m-%d")
            dias = (data_envio - data_entrada).days
        except:
            dias = "-"

        resultados.append({
            "id": d[0],
            "codigo": d[1],
            "entrada": d[2],
            "envio": d[3],
            "ensaios": ", ".join([d[4], d[5], d[6], d[7]]),
            "dias": dias
        })

    return render_template('lista.html', dados=resultados)


@app.route('/dashboard')
def dashboard():
    conn = sqlite3.connect(db_path)
    df = pd.read_sql_query("SELECT * FROM amostras", conn)
    conn.close()

    if df.empty:
        return render_template(
            'dashboard.html',
            total=0,
            media_geral=0,
            resumo=[]
        )

    df["dias"] = (
        pd.to_datetime(df["data_envio"], errors="coerce") -
        pd.to_datetime(df["data_entrada"], errors="coerce")
    ).dt.days

    total = len(df)
    media_geral = round(df["dias"].mean(), 1)

    ensaios = ["FQ - 03", "FQ - 04", "FQ - 07", "FQ - 12", "FQ - 18"]

    resumo = []

    for ensaio in ensaios:
        filtro = (
            (df["ensaio1"] == ensaio) |
            (df["ensaio2"] == ensaio) |
            (df["ensaio3"] == ensaio) |
            (df["ensaio4"] == ensaio)
        )

        df_ensaio = df[filtro]

        quantidade = len(df_ensaio)
        media = round(df_ensaio["dias"].mean(), 1) if quantidade > 0 else 0

        resumo.append({
            "ensaio": ensaio,
            "quantidade": quantidade,
            "media": media
        })

    return render_template(
        'dashboard.html',
        total=total,
        media_geral=media_geral,
        resumo=resumo
    )


@app.route('/exportar')
def exportar():
    criar_banco()

    if not os.path.exists(excel_path):
        flash("Arquivo Excel não encontrado. Verifique se 'analise de dados.xlsx' está na mesma pasta.")
        return redirect('/')

    conn = sqlite3.connect(db_path)
    df = pd.read_sql_query("SELECT * FROM amostras", conn)
    conn.close()

    df["dias"] = (
        pd.to_datetime(df["data_envio"], errors="coerce") -
        pd.to_datetime(df["data_entrada"], errors="coerce")
    ).dt.days

    from openpyxl import load_workbook

    wb = load_workbook(excel_path)

    if 'dados' not in wb.sheetnames:
        ws = wb.create_sheet('dados')
    else:
        ws = wb['dados']

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    for col_num, coluna in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_num, value=coluna)

    for row_num, linha in enumerate(df.itertuples(index=False), start=2):
        for col_num, valor in enumerate(linha, start=1):
            ws.cell(row=row_num, column=col_num, value=valor)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    wb.save(excel_path)

    flash("Dados exportados com sucesso para a planilha!")
    return redirect('/')


if __name__ == '__main__':
    app.run(debug=True)